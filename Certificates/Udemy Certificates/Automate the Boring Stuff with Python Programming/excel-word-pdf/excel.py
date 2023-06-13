import openpyxl;

path = "excel-word-pdf/docs/CHS-Chem.xlsx";
book = openpyxl.load_workbook(path);
sheet = book.get_sheet_by_name("Sheet1");
#str(Sheet["A1"].value);
#cell = sheet.cell(row = 1, column=2);
for i in range(4, 13):
    print(f"{i}:", sheet.cell(row = i, column = 1).value);